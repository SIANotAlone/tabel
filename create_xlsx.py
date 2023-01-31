from openpyxl import Workbook
from openpyxl import drawing
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
#from tqdm import tqdm

import os

class Create_xlsx:
    
    def __init__(self, output_folder):
        self.output_folder = output_folder

    


    def create(self, students):
        print("Generating xlsx documents...")
    
 
        literals = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)


        #rating
        ratings = {}
        for student in students:
            rating = 0
            for key,value in student["subjects"].items():
                try:
                    rating += int(value)
                except ValueError:
                    if type(value) == "<class 'str'>":
                        pass
            #print(student["second_semester"])
            for key,value in student["second_semester"].items():
                try:
                    rating += int(value)
                except ValueError:
                    if type(value) == "<class 'str'>":
                        pass
            ratings[student["name"]] = rating
        
        sort = sorted(ratings.items(), key=lambda x:x[1])
        converted_dict = dict(sort)
        
        t=25
        final_rating = {}
        for key,value in converted_dict.items():
            final_rating[key] = t
            t -=1
        #print(final_rating)


        #for student in tqdm(students):
        for student in students:

            wb = Workbook()
            ws = wb.active
            
            img = drawing.image.Image('tryzub.png')
            img.anchor = 'A1'
            ws.add_image(img)
            img1 = drawing.image.Image('tryzub.png')
            img1.anchor = 'G1'
            ws.add_image(img1)
         

            ws.column_dimensions['A'].width = 32
            rd = ws.row_dimensions[1] # get dimension for row 3
            rd.height = 45
            a1 = ws['A1']
            ft1 = Font(name='Montserrat', size=16, bold=True)
            a1.font = ft1
           
            a1.alignment = Alignment(horizontal="center", vertical="center")

            #font = Font(name='Montserrat',bold=True,size=16)
            #alignment=Alignment(horizontal='center')
            #ws.column_dimensions['A'] = alignment
            #ws.column_dimensions['B'].alignment = Alignment(horizontal='center')   - doesn`t work
            ws["A1"] = "Табель"
            ws["A2"] = "досягнень у навчанні та відвідування гімназії №86 учня "
            ws["A2"].alignment = Alignment(horizontal="right", vertical="center")
            ws["F2"] = student["class"]
            ws["F2"].alignment = Alignment(horizontal="center", vertical="center")
            ws["F2"].font = Font(name='Montserrat', size=12, bold=True)
            ws["G2"] = "класу"
            ws["G2"].alignment = Alignment(horizontal="center", vertical="center")
            ws["B4"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A3"] = "за"
            ws["A3"].alignment = Alignment(horizontal="right", vertical="center")
            ws["B3"] = student["years"]
            ws["B3"].alignment = Alignment(horizontal="center", vertical="center")
            ws["B3"].font = Font(name='Montserrat', size=12, bold=True)
            ws["D3"] = "навчальний рік"
            ws["D3"].alignment = Alignment(horizontal="left", vertical="center")
            ws["A4"] = "Прізвище та ім'я учня -"

            highlight1 = NamedStyle(name="highlight")
            bd1 = Side(style='thin', color="000000")
            highlight1.border = Border(left=bd1, top=bd1, right=bd1, bottom=bd1)

            ws["A5"] = "Предмети"
            
            ws["B5"] = "Бали"
            ws["B5"].style = highlight1
            ws["B5"].alignment = Alignment(horizontal="center", vertical="center")
            ws["B5"].font = Font(name='Montserrat', size=16)
            

            ws["A5"].style = highlight1
            ws["A5"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A5"].font = Font(name='Montserrat', size=16)
           

            ws["C5"].style = highlight1
            ws["D5"].style = highlight1
            ws["E5"].style = highlight1
            ws["F5"].style = highlight1
            ws["G5"].style = highlight1
            ws["B6"] = "за семестр"
            ws["B6"].style = highlight1
            ws["B6"].alignment = Alignment(horizontal="center", vertical="center")
            ws["B7"] = "I"
            ws["B7"].style = highlight1
            ws["B7"].alignment = Alignment(horizontal="center", vertical="center")
            ws["C7"] = "II"
            ws["C7"].style = highlight1
            ws["C7"].alignment = Alignment(horizontal="center", vertical="center")
            ws["D6"] = "Річні"
            ws["D6"].style = highlight1
            ws["D6"].alignment = Alignment(horizontal="center", vertical="center")
            ws["E6"] = "Навчальна практика"
            ws["E6"].style = highlight1
            ws["E6"].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            ws["E6"].font = Font(name='Montserrat', size=5)
            ws["F6"] = "Державна підсумкова атестація"
            ws["F6"].style = highlight1
            ws["F6"].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            ws["F6"].font = Font(name='Montserrat', size=5)
            ws["G6"] = "Підсумкова"
            ws["G6"].style = highlight1
            ws["G6"].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            ws["G6"].font = Font(name='Montserrat', size=5)   
           

            ws["B4"] = student["name"]
            ws["B4"].font = ft1
            
            
            
            i=8
            for key, value in student["subjects"].items():
                #print("key: " + str(key) + " value: " + str(value))
                if value != "":
                    ws["A"+str(i)] = key
                    
                    ws["B"+str(i)] = value
                    try:
                        ws["C"+str(i)] = student["second_semester"][key]
                        for_year =  (student["second_semester"][key] + value)//2
                        #print(for_year)
                        if (student["second_semester"][key] + value)%2 > 0.5:
                            ws["D"+str(i)] = for_year +1
                        else:
                            ws["D"+str(i)] = for_year
                        if key == "Рейтінгове місце з 25":
                            ws["D"+str(i)] = final_rating[student["name"]]
                            #ws["D"+str(i)] = ""
                        if (student["second_semester"][key] == 0):
                            ws["C"+str(i)] = ""
                            ws["C"+str(i-1)] = ""
                            ws["D"+str(i)] = ""
                            ws["D"+str(i-1)] = ""
                    except KeyError:
                        pass
                    except TypeError:
                        pass
                    
                    ws["A"+str(i)].style = highlight1
                    ws["B"+str(i)].style = highlight1
                    ws["C"+str(i)].style = highlight1
                    ws["D"+str(i)].style = highlight1
                    ws["E"+str(i)].style = highlight1
                    ws["F"+str(i)].style = highlight1
                    ws["G"+str(i)].style = highlight1
                    ws["A"+str(i)].font = Font(name='Montserrat')
                    ws["B"+str(i)].font = Font(name='Montserrat')
                    ws["C"+str(i)].font = Font(name='Montserrat')
                    ws["D"+str(i)].font = Font(name='Montserrat')
                    grade = ws["B"+str(i)]
                    grade.alignment = Alignment(horizontal="center", vertical="center")
                    grade = ws["C"+str(i)]
                    grade.alignment = Alignment(horizontal="center", vertical="center")
                    grade = ws["D"+str(i)]
                    grade.alignment = Alignment(horizontal="center", vertical="center")
                    i+=1
                    
           
            ws["A"+str(i+1)] = 'Підпис класного керівника'
            ws["A"+str(i+2)] = 'Підпис батьків'
            ws["F"+str(i+1)] = '_________________'
            ws["F"+str(i+2)] = '_________________'
            ##### Borders
            # highlight = NamedStyle(name="highlight")
            # bd = Side(style='thick', color="000000")
            # highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
            # ws['A1'].style = highlight
            
            ws.merge_cells('A1:G1')
            ws.merge_cells('A2:E2')
            ws.merge_cells('B3:C3')
            ws.merge_cells('D3:G3')
            ws.merge_cells('A5:A7')
            ws.merge_cells('B5:G5')
            ws.merge_cells('B6:C6')
            ws.merge_cells('D6:D7')
            ws.merge_cells('E6:E7')
            ws.merge_cells('F6:F7')
            ws.merge_cells('G6:G7')
            ws.merge_cells('B4:G4')
            wb.save(self.output_folder + student["name"]+"_табель"+".xlsx")
            name = student["name"]
            print(f"Табель {name} збережено.")
        print(f"Всі табеля успішно згенеровані в теці {self.output_folder}")
