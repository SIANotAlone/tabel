from openpyxl import Workbook
from openpyxl import drawing
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
#from tqdm import tqdm

import os


class Create_xlsx_statistics:
    def __init__(self, output_folder) -> None:
        self.output_folder = output_folder
        
    def create(self, data, semester, klas, f_semester, s_semester, year, f_teacher, s_teacher):
        print("Generating xlsx report...")

        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)
    
        wb = Workbook()
     
        
        ws = wb.active
        
        
        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 7
        ws.column_dimensions['D'].width = 4
        ws.column_dimensions['E'].width = 4
        ws.column_dimensions['F'].width = 4
        ws.column_dimensions['G'].width = 7
        ws.column_dimensions['H'].width = 7
        ws.column_dimensions['I'].width = 4
        ws.column_dimensions['J'].width = 4
        ws.column_dimensions['K'].width = 4
        ws.column_dimensions['L'].width = 7
        ws.column_dimensions['M'].width = 7
        ws.column_dimensions['N'].width = 4
        ws.column_dimensions['O'].width = 4
        ws.column_dimensions['P'].width = 4
        ws.column_dimensions['Q'].width = 7
        ws.column_dimensions['R'].width = 7
        ws.column_dimensions['S'].width = 4
        ws.column_dimensions['T'].width = 4
        ws.column_dimensions['U'].width = 4
        ws.column_dimensions['V'].width = 7
        ws.column_dimensions['W'].width = 7
        ws.column_dimensions['X'].width = 12

        ws.row_dimensions[3].height = 30



        ws["B1"] = "Якісні показники навчальних досягнень  учнів"
        ws["G1"] = klas
        if semester == "I семестр":
            ws["H1"] = f_semester
            ws["V1"] = f_teacher
        else:
            ws["H1"] = s_semester
            ws["V1"] = s_teacher

        ws["L1"] = year
        ws["N1"] = "н.р."
        ws["Q1"] = "Класний керівник:__________"

        ws["A2"] = "№ п/п"
        ws["B2"] = "Предмет"
        ws["C2"] = "Кількість атестованих учнів"

        ws["D2"] = "І рівень(початковий)"
        ws["I2"] = "II рівень(середній)"
        ws["N2"] = "ІІІ рівень(достатній)"
        ws["S2"] = "ІV рівень(високий)"

        ws["X2"] = "Cередній бал по предмету"

        ws["D3"] = "1"
        ws["E3"] = "2"
        ws["F3"] = "3"
        ws["G3"] = "Всього"
        ws["H3"] = "%"

        ws["I3"] = "4"
        ws["J3"] = "5"
        ws["K3"] = "6"
        ws["L3"] = "Всього"
        ws["M3"] = "%"

        ws["N3"] = "7"
        ws["O3"] = "8"
        ws["P3"] = "9"
        ws["Q3"] = "Всього"
        ws["R3"] = "%"

        ws["S3"] = "10"
        ws["T3"] = "11"
        ws["U3"] = "12"
        ws["V3"] = "Всього"
        ws["W3"] = "%"

        
        #ws["E6"].font = Font(name='Montserrat', size=5)

        highlight1 = NamedStyle(name="highlight")
        bd1 = Side(style='thin', color="000000")
        highlight1.border = Border(left=bd1, top=bd1, right=bd1, bottom=bd1)

        ws["D2"].style = highlight1
        ws["I2"].style = highlight1
        ws["N2"].style = highlight1
        ws["S2"].style = highlight1
        ws["X2"].style = highlight1

        ws["A2"].style = highlight1
        ws["B2"].style = highlight1
        ws["C2"].style = highlight1
        ws["D3"].style = highlight1
        ws["E3"].style = highlight1
        ws["F3"].style = highlight1
        ws["G3"].style = highlight1
        ws["H3"].style = highlight1
        ws["I3"].style = highlight1
        ws["J3"].style = highlight1
        ws["K3"].style = highlight1
        ws["L3"].style = highlight1
        ws["M3"].style = highlight1
        ws["N3"].style = highlight1
        ws["O3"].style = highlight1
        ws["P3"].style = highlight1
        ws["Q3"].style = highlight1
        ws["R3"].style = highlight1
        ws["S3"].style = highlight1
        ws["T3"].style = highlight1
        ws["U3"].style = highlight1
        ws["V3"].style = highlight1
        ws["W3"].style = highlight1
        ws["X3"].style = highlight1
        
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["C2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["D3"].alignment = Alignment(horizontal="center", vertical="center")
        ws["E3"].alignment = Alignment(horizontal="center", vertical="center")
        ws["F3"].alignment = Alignment(horizontal="center", vertical="center")
        ws["G3"].alignment = Alignment(horizontal="center", vertical="center")
        ws["H3"].alignment = Alignment(horizontal="center", vertical="center")
        ws["I3"].alignment = Alignment(horizontal="center", vertical="center")
        ws["J3"].alignment = Alignment(horizontal="center", vertical="center")
        ws["K3"].alignment = Alignment(horizontal="center", vertical="center")
        ws["L3"].alignment = Alignment(horizontal="center", vertical="center")
        ws["M3"].alignment = Alignment(horizontal="center", vertical="center")
        ws["N3"].alignment = Alignment(horizontal="center", vertical="center")
        ws["O3"].alignment = Alignment(horizontal="center", vertical="center")
        ws["P3"].alignment = Alignment(horizontal="center", vertical="center")
        ws["Q3"].alignment = Alignment(horizontal="center", vertical="center")
        ws["R3"].alignment = Alignment(horizontal="center", vertical="center")
        ws["S3"].alignment = Alignment(horizontal="center", vertical="center")
        ws["T3"].alignment = Alignment(horizontal="center", vertical="center")
        ws["U3"].alignment = Alignment(horizontal="center", vertical="center")
        ws["V3"].alignment = Alignment(horizontal="center", vertical="center")
        ws["W3"].alignment = Alignment(horizontal="center", vertical="center")
        ws["X3"].alignment = Alignment(horizontal="center", vertical="center")

        ws["D2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["I2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["N2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["S2"].alignment = Alignment(horizontal="center", vertical="center")


        ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        ws["C2"].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        ws["X2"].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
        ws["C2"].font = Font(size=8)
        class_average_all_subjects =[]
        i=1
        for key,value in data.items():
           
            try:
                ws["G"+str(i+3)] = value["level1_count"]
                ws["H"+str(i+3)] = str(value["level1_percent"]) + "%"
                ws["L"+str(i+3)] = value["level2_count"]
                ws["M"+str(i+3)] = str(value["level2_percent"])+ "%"
                ws["Q"+str(i+3)] = value["level3_count"]
                ws["R"+str(i+3)] = str(value["level3_percent"])+ "%"
                ws["V"+str(i+3)] = value["level4_count"]
                ws["W"+str(i+3)] = str(value["level4_percent"])+ "%"
                ws["X"+str(i+3)] = value["average"]
                ws["C"+str(i+3)] = value["sum"]
                class_average_all_subjects.append(value["average"])

                ws["A"+str(i+3)] = str(i)
                ws["B"+str(i+3)] = key
                

                ws["D"+str(i+3)] = value["1"]
                ws["E"+str(i+3)] = value["2"]
                ws["F"+str(i+3)] = value["3"]

                ws["I"+str(i+3)] = value["4"]
                ws["J"+str(i+3)] = value["5"]
                ws["K"+str(i+3)] = value["6"]

                ws["N"+str(i+3)] = value["7"]
                ws["O"+str(i+3)] = value["8"]
                ws["P"+str(i+3)] = value["9"]

                ws["S"+str(i+3)] = value["10"]
                ws["T"+str(i+3)] = value["11"]
                ws["U"+str(i+3)] = value["12"]
                


                ws["A"+str(i+3)].style = highlight1
                ws["B"+str(i+3)].style = highlight1
                ws["D"+str(i+3)].style = highlight1
                ws["E"+str(i+3)].style = highlight1
                ws["F"+str(i+3)].style = highlight1
                ws["I"+str(i+3)].style = highlight1
                ws["J"+str(i+3)].style = highlight1
                ws["K"+str(i+3)].style = highlight1
                ws["N"+str(i+3)].style = highlight1
                ws["O"+str(i+3)].style = highlight1
                ws["P"+str(i+3)].style = highlight1
                ws["S"+str(i+3)].style = highlight1
                ws["T"+str(i+3)].style = highlight1
                ws["U"+str(i+3)].style = highlight1

                ws["C"+str(i+3)].style = highlight1
                ws["G"+str(i+3)].style = highlight1
                ws["H"+str(i+3)].style = highlight1
                ws["L"+str(i+3)].style = highlight1
                ws["M"+str(i+3)].style = highlight1
                ws["Q"+str(i+3)].style = highlight1
                ws["R"+str(i+3)].style = highlight1
                ws["V"+str(i+3)].style = highlight1
                ws["W"+str(i+3)].style = highlight1
                ws["X"+str(i+3)].style = highlight1

                ws["C"+str(i+3)].alignment = Alignment(horizontal="center", vertical="center")
                ws["D"+str(i+3)].alignment = Alignment(horizontal="center", vertical="center")
                ws["E"+str(i+3)].alignment = Alignment(horizontal="center", vertical="center")
                ws["F"+str(i+3)].alignment = Alignment(horizontal="center", vertical="center")
                ws["G"+str(i+3)].alignment = Alignment(horizontal="center", vertical="center")
                ws["H"+str(i+3)].alignment = Alignment(horizontal="center", vertical="center")
                ws["I"+str(i+3)].alignment = Alignment(horizontal="center", vertical="center")
                ws["J"+str(i+3)].alignment = Alignment(horizontal="center", vertical="center")
                ws["K"+str(i+3)].alignment = Alignment(horizontal="center", vertical="center")
                ws["L"+str(i+3)].alignment = Alignment(horizontal="center", vertical="center")
                ws["M"+str(i+3)].alignment = Alignment(horizontal="center", vertical="center")
                ws["N"+str(i+3)].alignment = Alignment(horizontal="center", vertical="center")
                ws["O"+str(i+3)].alignment = Alignment(horizontal="center", vertical="center")
                ws["P"+str(i+3)].alignment = Alignment(horizontal="center", vertical="center")
                ws["Q"+str(i+3)].alignment = Alignment(horizontal="center", vertical="center")
                ws["R"+str(i+3)].alignment = Alignment(horizontal="center", vertical="center")
                ws["S"+str(i+3)].alignment = Alignment(horizontal="center", vertical="center")
                ws["T"+str(i+3)].alignment = Alignment(horizontal="center", vertical="center")
                ws["U"+str(i+3)].alignment = Alignment(horizontal="center", vertical="center")
                ws["V"+str(i+3)].alignment = Alignment(horizontal="center", vertical="center")
                ws["W"+str(i+3)].alignment = Alignment(horizontal="center", vertical="center")
                ws["X"+str(i+3)].alignment = Alignment(horizontal="center", vertical="center")



                i+=1
            except KeyError:
                pass
        class_average = 0

        for grd in class_average_all_subjects:
            class_average+=grd
        class_average = class_average/len(class_average_all_subjects)

        ws["B"+str(i+5)] = "Експерт_______________________________"
        ws["W"+str(i+3)] = "клас"
        ws["X"+str(i+3)] = "{:.2f}".format(class_average)
        #ws["X"+str(i+4)].style = highlight1
        ws["X"+str(i+3)].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells('A2:A3')
        ws.merge_cells('B2:B3')
        ws.merge_cells('C2:C3')
        ws.merge_cells('D2:H2')
        ws.merge_cells('I2:M2')
        ws.merge_cells('N2:R2')
        ws.merge_cells('S2:W2')
        ws.merge_cells('X2:X3')




        wb.save(self.output_folder + "звіт_за_"+semester +".xlsx")
        print("Report successfully generated in current folder.")

